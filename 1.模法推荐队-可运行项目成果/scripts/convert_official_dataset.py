#!/usr/bin/env python3
"""
convert_official_dataset.py - Convert official Excel dataset to JSONL format.

Reads: data/raw/official/official_dataset.xlsx
Generates:
  - data/official/questions_train.jsonl
  - data/official/questions_test.jsonl
  - data/official/questions_val.jsonl
  - data/official/questions_all.jsonl
  - data/official/model_catalog_raw.jsonl
  - data/official/model_catalog_structured.jsonl
  - data/official/model_name_map.json
  - data/eval_official/intent_eval_official.jsonl
  - data/eval_official/tag_eval_official.jsonl
  - data/eval_official/topk_eval_official.jsonl
  - data/eval_official/combo_eval_official_manual.jsonl

Usage:
    python scripts/convert_official_dataset.py
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def classify_domain(model_name: str) -> str:
    """Classify the business domain from model name."""
    risk_keywords = [
        '风控', '风险', '欺诈', '反诈', '违约', '逾期', '催收', '评分卡', 'A卡', 'B卡',
        '压力测试', '合规监测', '非法集资', '贷款中介', '垒小户', '准入', '黑名单', '白名单', '信用'
    ]
    mkt_keywords = [
        '营销', '推荐', '促活', '拓客', '增存', '流失预警', '流失预测', '挖掘', '转化',
        '响应', '分期', '额度调整', '交叉销售', '理财', '保险', '财富', '存款', '客户经营', '精准营销', '回捞'
    ]
    ops_keywords = [
        '运营', '管理', '监测', '监控', '分析', '评价', '测试', '效率', '资源', '排班',
        '网点', '流程', '合规', '报表', '审计', '检查'
    ]

    risk_score = sum(1 for k in risk_keywords if k in model_name)
    mkt_score = sum(1 for k in mkt_keywords if k in model_name)
    ops_score = sum(1 for k in ops_keywords if k in model_name)

    scores = [(risk_score, 'credit_risk'), (mkt_score, 'customer_marketing'), (ops_score, 'operation_management')]
    scores.sort(reverse=True)
    return scores[0][1]


def classify_task(model_name: str, query: str) -> str:
    """Classify the specific task from model name and query."""
    task_keywords = {
        'fraud_detection': ['欺诈', '反诈', '骗贷', '反欺诈', '异常检测'],
        'admission_scoring': ['准入', '评分卡', '准入评分', '信用评分'],
        'default_prediction': ['违约', '逾期', 'B卡', '违约概率'],
        'marketing_targeting': ['营销', '精准营销', '目标客户', '客群筛选'],
        'conversion_prediction': ['转化', '促活', '激活', '首刷', '首贷'],
        'churn_prediction': ['流失', '流失预警', '流失预测', '挽留'],
        'value_assessment': ['价值', '分层', '评估', '高价值'],
        'cross_selling': ['交叉销售', '联动', '推荐', '产品推荐'],
        'early_warning': ['预警', '监测', '监控', 'A卡'],
        'collection_optimization': ['催收', '清收', '回收'],
        'amount_calculation': ['额度', '授信', '限额'],
        'customer_segmentation': ['分层', '分群', '客群', '分类'],
        'risk_network_analysis': ['关联', '担保圈', '团伙', '中介'],
        'aml_compliance': ['反洗钱', '洗钱', '合规', '监管'],
        'process_optimization': ['流程', '优化', '效率', '改进'],
        'resource_planning': ['排班', '资源配置', '人力', '规划'],
        'demand_forecasting': ['预测', '客流', '业务量', '趋势'],
        'performance_analysis': ['绩效', '效能', '评价', '考核'],
    }

    combined = model_name + ' ' + query
    scores = []
    for task, keywords in task_keywords.items():
        score = sum(1 for k in keywords if k in combined)
        if score > 0:
            scores.append((score, task))
    scores.sort(reverse=True)
    return scores[0][1] if scores else 'model_recommendation'


def extract_tags(query: str, model_name: str, domain: str, task: str) -> list:
    """Weak annotation: extract expected tags from content."""
    tags = set()

    domain_map = {
        'credit_risk': '信贷风控',
        'customer_marketing': '客户营销',
        'operation_management': '运营管理'
    }
    tags.add(domain_map.get(domain, domain))

    task_tag_map = {
        'fraud_detection': ['反欺诈', '欺诈识别'],
        'admission_scoring': ['准入评分', '评分卡'],
        'default_prediction': ['违约预测', '逾期预测'],
        'marketing_targeting': ['精准营销', '营销转化'],
        'conversion_prediction': ['转化预测', '响应预测'],
        'churn_prediction': ['流失预警', '流失预测'],
        'value_assessment': ['价值评估', '客户分层'],
        'cross_selling': ['交叉销售', '产品推荐'],
        'early_warning': ['预警监控', '风险预警'],
        'collection_optimization': ['催收排序', '清收'],
        'amount_calculation': ['额度测算', '授信'],
        'customer_segmentation': ['客群分层', '客户分群'],
        'risk_network_analysis': ['关联分析', '担保圈'],
        'aml_compliance': ['反洗钱', '合规管理'],
        'process_optimization': ['流程优化', '效率提升'],
        'resource_planning': ['资源优化', '排班'],
        'demand_forecasting': ['需求预测', '业务量预测'],
        'performance_analysis': ['绩效分析', '效能评价'],
    }
    task_tags = task_tag_map.get(task, [])
    if task_tags:
        tags.add(task_tags[0])

    content_tags = {
        '农户': '农户', '小微': '小微企业和个体工商户', '个体工商户': '小微企业和个体工商户',
        '对公': '对公客户/企业', '企业': '对公客户/企业',
        '信用卡': '信用卡', '贷记卡': '信用卡',
        '消费贷': '消费贷', '房贷': '房贷', '存款': '存款',
        '理财': '理财产品', '保险': '保险', '手机银行': '手机银行',
        '收单': '收单商户', '商户': '收单商户', '县域': '县域客群',
        '新客': '新客', '存量': '存量客户', '高价值': '高净值客户',
        '农户小额': '小额贷款', '小额': '小额贷款', '首贷': '首贷',
        '涉农': '涉农贷款', '绿色': '绿色金融', '养老': '养老金融',
    }

    combined = query + model_name
    for keyword, tag in content_tags.items():
        if keyword in combined:
            tags.add(tag)

    return sorted(list(tags))


def build_combo_cases(name_to_id: dict) -> list:
    """Build manual combination evaluation cases."""
    return [
        {
            "case_id": "COMBO_OFFICIAL_001",
            "name": "农户小额贷款贷前全流程风控",
            "description": "农户申请小额贷款，需要完成反欺诈检测、准入评分和额度测算",
            "query": "农户申请小额贷款，先做反欺诈检查，再评估准入资格，最后测算合理额度",
            "gold_model_ids": [name_to_id.get("阳光E贷贷前准入模型", ""), name_to_id.get("垒小户准入评分卡", ""), name_to_id.get("垒小户违约概率模型", "")],
            "gold_model_names": ["阳光E贷贷前准入模型", "垒小户准入评分卡", "垒小户违约概率模型"],
            "scenario": "credit_risk",
            "complexity": "high"
        },
        {
            "case_id": "COMBO_OFFICIAL_002",
            "name": "小微企业贷款风控与营销组合",
            "description": "小微企业客户既有贷款风控需求，也有产品营销机会",
            "query": "对小微企业客户做贷前风险评估，同时推荐合适的贷款产品",
            "gold_model_ids": [name_to_id.get("经营贷款贷前准入模型", ""), name_to_id.get("收单商户贷款营销模型", "")],
            "gold_model_names": ["经营贷款贷前准入模型", "收单商户贷款营销模型"],
            "scenario": "cross_domain",
            "complexity": "medium"
        },
        {
            "case_id": "COMBO_OFFICIAL_003",
            "name": "信用卡客户全生命周期经营",
            "description": "信用卡客户从贷前申请到贷后营销的全流程",
            "query": "信用卡客户贷前反欺诈、准入评分，贷后账单分期营销和额度调整",
            "gold_model_ids": [name_to_id.get("信用卡贷前反欺诈机器学习模型", ""), name_to_id.get("信用卡贷前申请评分模型_有征信", ""), name_to_id.get("贷记卡账单分期营销模型", ""), name_to_id.get("贷记卡固定额度调整模型", "")],
            "gold_model_names": ["信用卡贷前反欺诈机器学习模型", "信用卡贷前申请评分模型_有征信", "贷记卡账单分期营销模型", "贷记卡固定额度调整模型"],
            "scenario": "cross_domain",
            "complexity": "high"
        },
        {
            "case_id": "COMBO_OFFICIAL_004",
            "name": "对公客户风险与价值综合评估",
            "description": "对公客户的价值评估与风险监控组合",
            "query": "评估对公客户的综合价值和风险水平，识别高价值低风险客户",
            "gold_model_ids": [name_to_id.get("对公客户价值与风险评分模型", "")],
            "gold_model_names": ["对公客户价值与风险评分模型"],
            "scenario": "credit_risk",
            "complexity": "medium"
        },
        {
            "case_id": "COMBO_OFFICIAL_005",
            "name": "县域新客首贷营销转化",
            "description": "识别县域新客中的首贷潜力客户并做精准营销",
            "query": "筛选县域新客中可能成为首贷户的客户，并预测其转化概率",
            "gold_model_ids": [name_to_id.get("潜在高价值客户挖掘模型", ""), name_to_id.get("潜客挖掘消费贷营销模型", "")],
            "gold_model_names": ["潜在高价值客户挖掘模型", "潜客挖掘消费贷营销模型"],
            "scenario": "customer_marketing",
            "complexity": "medium"
        },
        {
            "case_id": "COMBO_OFFICIAL_006",
            "name": "存量客户流失预警与挽回",
            "description": "预测存量客户流失风险并制定挽回策略",
            "query": "预测哪些存量客户可能流失，并推荐挽留措施",
            "gold_model_ids": [name_to_id.get("AUM10万以上客户流失预警模型", ""), name_to_id.get("高价值存款客户流失预测模型", ""), name_to_id.get("个贷流失回捞算法模型", "")],
            "gold_model_names": ["AUM10万以上客户流失预警模型", "高价值存款客户流失预测模型", "个贷流失回捞算法模型"],
            "scenario": "customer_marketing",
            "complexity": "high"
        },
        {
            "case_id": "COMBO_OFFICIAL_007",
            "name": "个贷全流程风控组合",
            "description": "个人贷款从贷前反欺诈到贷后催收的全流程",
            "query": "个人贷款申请时做反欺诈和准入评估，贷后逾期客户做催收排序",
            "gold_model_ids": [name_to_id.get("个贷全流程反欺诈机器学习模型", ""), name_to_id.get("信用卡贷前申请评分模型_有征信", ""), name_to_id.get("个贷逾期客群催收模型", "")],
            "gold_model_names": ["个贷全流程反欺诈机器学习模型", "信用卡贷前申请评分模型_有征信", "个贷逾期客群催收模型"],
            "scenario": "credit_risk",
            "complexity": "high"
        },
        {
            "case_id": "COMBO_OFFICIAL_008",
            "name": "收单商户综合经营",
            "description": "收单商户的反欺诈、价值分层、流失预测和贷款营销",
            "query": "收单商户做反欺诈监测、价值分层、流失预测和贷款营销",
            "gold_model_ids": [name_to_id.get("收单商户交易反欺诈模型", ""), name_to_id.get("收单商户价值分层及预授信模型", ""), name_to_id.get("收单商户流失预测模型", ""), name_to_id.get("收单商户贷款营销模型", "")],
            "gold_model_names": ["收单商户交易反欺诈模型", "收单商户价值分层及预授信模型", "收单商户流失预测模型", "收单商户贷款营销模型"],
            "scenario": "cross_domain",
            "complexity": "high"
        },
        {
            "case_id": "COMBO_OFFICIAL_009",
            "name": "开门红个贷营销回捞",
            "description": "开门红期间对个贷客户进行精准营销回捞",
            "query": "开门红期间对流失个贷客户做回捞营销",
            "gold_model_ids": [name_to_id.get("开门红个贷回捞精准营销模型", ""), name_to_id.get("个贷流失回捞算法模型", "")],
            "gold_model_names": ["开门红个贷回捞精准营销模型", "个贷流失回捞算法模型"],
            "scenario": "customer_marketing",
            "complexity": "medium"
        },
        {
            "case_id": "COMBO_OFFICIAL_010",
            "name": "房贷全流程管理",
            "description": "房贷客户从贷前审批到贷后预警的全流程",
            "query": "房贷客户贷前风险评估、违约概率预测和贷后预警监控",
            "gold_model_ids": [name_to_id.get("房贷用户按揭A卡模型", ""), name_to_id.get("房贷按揭违约概率模型", ""), name_to_id.get("房地产压力测试", "")],
            "gold_model_names": ["房贷用户按揭A卡模型", "房贷按揭违约概率模型", "房地产压力测试"],
            "scenario": "credit_risk",
            "complexity": "high"
        }
    ]


def main():
    BASE_DIR = Path(__file__).resolve().parent.parent
    EXCEL_PATH = BASE_DIR / "data" / "raw" / "official" / "official_dataset.xlsx"

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    official_dir = BASE_DIR / "data" / "official"
    official_dir.mkdir(parents=True, exist_ok=True)
    eval_official_dir = BASE_DIR / "data" / "eval_official"
    eval_official_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Read questions from all splits
    questions = []
    for sheet_name in ["训练集", "测试集", "验证集"]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if row and row[0]:
                questions.append({
                    "question_id": row[0],
                    "user_query": row[1],
                    "gold_model_name": row[2],
                    "split": row[3] if len(row) > 3 else sheet_name
                })

    # Read model catalog
    ws_catalog = wb["模型清单_参考"]
    catalog_rows = list(ws_catalog.iter_rows(values_only=True))
    catalog_map = {}
    duplicates = []
    for row in catalog_rows[1:]:
        if row and row[1]:
            name = row[1].strip()
            desc = row[2] if len(row) > 2 and row[2] else ""
            if name in catalog_map:
                duplicates.append(name)
            else:
                catalog_map[name] = desc

    # Create stable OFFICIAL_xxx IDs
    sorted_names = sorted(catalog_map.keys())
    name_to_id = {}
    id_to_name = {}
    for i, name in enumerate(sorted_names, 1):
        oid = f"OFFICIAL_{i:03d}"
        name_to_id[name] = oid
        id_to_name[oid] = name

    # Annotate questions with weak labels
    annotated = []
    for q in questions:
        model_name = q['gold_model_name']
        query = q['user_query']
        domain = classify_domain(model_name)
        task = classify_task(model_name, query)
        tags = extract_tags(query, model_name, domain, task)
        annotated.append({
            "question_id": q['question_id'],
            "user_query": query,
            "gold_model_id": name_to_id.get(model_name, ""),
            "gold_model_name": model_name,
            "split": q['split'],
            "intent_primary": "model_recommendation",
            "intent_domain": domain,
            "intent_task": task,
            "expected_tags": tags,
            "source": "official",
            "annotation_version": "1.0_weak",
            "needs_review": True
        })

    # Write question files
    splits = {k: [q for q in annotated if q['split'] == k] for k in ['train', 'test', 'val']}
    for split_name, qs in splits.items():
        path = official_dir / f"questions_{split_name}.jsonl"
        with open(path, 'w', encoding='utf-8') as f:
            for q in qs:
                f.write(json.dumps(q, ensure_ascii=False) + '\n')
        print(f"Written {path}: {len(qs)} questions")

    with open(official_dir / "questions_all.jsonl", 'w', encoding='utf-8') as f:
        for q in annotated:
            f.write(json.dumps(q, ensure_ascii=False) + '\n')
    print(f"Written questions_all.jsonl: {len(annotated)} questions")

    # Write model catalog files
    raw_catalog = [{"model_id": f"OFFICIAL_{i:03d}", "canonical_name": name, "description": catalog_map[name], "source": "official"}
                   for i, name in enumerate(sorted_names, 1)]
    with open(official_dir / "model_catalog_raw.jsonl", 'w', encoding='utf-8') as f:
        for item in raw_catalog:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')
    print(f"Written model_catalog_raw.jsonl: {len(raw_catalog)} models")

    structured = []
    for i, name in enumerate(sorted_names, 1):
        oid = f"OFFICIAL_{i:03d}"
        domain = classify_domain(name)
        total_q = sum(1 for q in questions if q['gold_model_name'] == name)
        structured.append({
            "model_id": oid,
            "canonical_name": name,
            "aliases": [],
            "domain": domain,
            "description": catalog_map[name],
            "source": "official",
            "total_questions": total_q
        })
    with open(official_dir / "model_catalog_structured.jsonl", 'w', encoding='utf-8') as f:
        for item in structured:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')
    print(f"Written model_catalog_structured.jsonl: {len(structured)} models")

    # Write name map
    with open(official_dir / "model_name_map.json", 'w', encoding='utf-8') as f:
        json.dump({
            "model_name_to_id": name_to_id,
            "id_to_name": id_to_name,
            "duplicate_names": duplicates,
            "total_unique": len(sorted_names),
            "total_with_duplicates": len(catalog_rows) - 1
        }, f, ensure_ascii=False, indent=2)
    print("Written model_name_map.json")

    # Write evaluation files
    intent_eval = []
    tag_eval = []
    topk_eval = []
    for q in annotated:
        intent_eval.append({
            "test_id": q['question_id'],
            "query": q['user_query'],
            "expected_intent": q['intent_primary'],
            "expected_domain": q['intent_domain'],
            "expected_task": q['intent_task'],
            "difficulty": "medium"
        })
        tag_eval.append({
            "test_id": q['question_id'],
            "query": q['user_query'],
            "expected_tags": q['expected_tags'],
            "difficulty": "medium"
        })
        topk_eval.append({
            "test_id": q['question_id'],
            "query": q['user_query'],
            "gold_model_id": q['gold_model_id'],
            "gold_model_name": q['gold_model_name'],
            "expected_model_ids": [q['gold_model_id']],
            "k": 5,
            "scenario": q['intent_domain']
        })

    for fname, data in [
        ("intent_eval_official.jsonl", intent_eval),
        ("tag_eval_official.jsonl", tag_eval),
        ("topk_eval_official.jsonl", topk_eval),
    ]:
        with open(eval_official_dir / fname, 'w', encoding='utf-8') as f:
            for item in data:
                f.write(json.dumps(item, ensure_ascii=False) + '\n')
        print(f"Written {fname}: {len(data)} items")

    # Write combo eval
    combo_cases = build_combo_cases(name_to_id)
    with open(eval_official_dir / "combo_eval_official_manual.jsonl", 'w', encoding='utf-8') as f:
        for item in combo_cases:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')
    print(f"Written combo_eval_official_manual.jsonl: {len(combo_cases)} cases")

    print("\n=== CONVERSION COMPLETE ===")
    print(f"Train: {len(splits['train'])}, Test: {len(splits['test'])}, Val: {len(splits['val'])}")
    print(f"Total: {len(annotated)}, Unique models: {len(sorted_names)}")
    print(f"Duplicates: {duplicates}")

    # Verify overlap with demo models
    demo_names = set()
    for f in (BASE_DIR / "data" / "knowledge").glob("*.json"):
        if f.name[0:3] in ("RIS", "MKT", "OPS"):
            data = json.load(open(f, 'r', encoding='utf-8'))
            demo_names.add(data.get('model_name', ''))
    overlap = set(sorted_names) & demo_names
    print(f"Demo/Official exact overlap: {len(overlap)} (expected: 0)")
    if overlap:
        print(f"WARNING: Overlapping names found: {overlap}")


if __name__ == "__main__":
    main()
