#!/usr/bin/env python3
"""
prepare_official_dataset.py

从官方 zip 读取真实 Excel，生成 data/official_60/ 和 reports/official_dataset/ 的全部官方数据产物。

输出清单:
  data/raw/official_dataset/大模型驱动的模型市场智能推荐助手研究_数据集.xlsx
  data/official_60/models.jsonl
  data/official_60/models_official_60.json
  data/official_60/query_model_eval.jsonl
  data/official_60/queries_train.jsonl
  data/official_60/queries_test.jsonl
  data/official_60/queries_val.jsonl
  data/official_60/dataset_manifest.json
  reports/official_dataset/official_excel_inspection.md
  reports/official_dataset/official_excel_inspection.json
  reports/official_dataset/duplicate_models.md
  reports/official_dataset/duplicate_models.json
  reports/official_dataset/model_question_distribution.md
  reports/official_dataset/model_question_distribution.json

用法:
  python scripts/prepare_official_dataset.py
"""

import json
import os
import shutil
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


# ── 路径 ──────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent

ZIP_PATH = BASE_DIR / (
    "16-知识图谱与智能推荐赛道-江苏农商联合银行"
    "-大模型驱动的模型市场智能推荐助手研究.zip"
)

RAW_DIR = BASE_DIR / "data" / "raw" / "official_dataset"
OFFICIAL_DIR = BASE_DIR / "data" / "official_60"
REPORT_DIR = BASE_DIR / "reports" / "official_dataset"

EXCEL_FILENAME = "大模型驱动的模型市场智能推荐助手研究_数据集.xlsx"
EXCEL_PATH = RAW_DIR / EXCEL_FILENAME


def project_relative(path: Path) -> str:
    """Return a stable repository-relative path for generated metadata."""
    try:
        return path.resolve().relative_to(BASE_DIR.resolve()).as_posix()
    except ValueError:
        return str(path)

# ── Sheet 名称 ────────────────────────────────────────────────────────────
SHEET_TRAIN = "训练集"
SHEET_TEST = "测试集"
SHEET_VAL = "验证集"
SHEET_ALL = "全部数据"
SHEET_MODELS = "模型清单_参考"
SHEET_STATS = "统计信息"

# ── 域映射（关键词 → domain） ──────────────────────────────────────────────
DOMAIN_KEYWORDS: list[tuple[list[str], str]] = [
    (["风险", "风控", "违约", "欺诈", "评分卡", "准入", "逾期", "反洗钱", "合规"], "credit_risk"),
    (["营销", "推荐", "促活", "流失", "客户", "分期", "首贷", "促", "激励"], "customer_marketing"),
    (["运营", "管理", "预警", "监控", "识别", "运维", "流程", "分层"], "operation_management"),
]

DOMAIN_FALLBACK = "operation_management"

# ── 中文停用词（用于 tag 提取） ────────────────────────────────────────────
STOP_WORDS: set[str] = {
    "模型", "业务", "客户", "数据", "分析", "系统", "管理", "产品", "服务",
    "方案", "平台", "场景", "工具", "能力", "识别", "预测", "评估", "策略",
    "流程", "指标", "报告", "项目", "需求", "应用", "方式", "方法", "技术",
    "模型", "的", "和", "与", "及", "或", "在", "对", "为", "以", "等",
    "基于", "进行", "通过", "利用", "实现", "提供", "支持", "包括",
}


def _ensure_dir(d: Path) -> None:
    """确保目录存在，若已存在则清空（幂等辅助）。"""
    d.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════
# 1. 解压
# ═══════════════════════════════════════════════════════════════════════════
def extract_excel() -> Path:
    """Return the official Excel file, extracting from local zip only if needed."""
    if EXCEL_PATH.exists():
        return EXCEL_PATH

    if not ZIP_PATH.exists():
        raise FileNotFoundError(
            "Official Excel not found. Expected committed Excel at "
            f"{EXCEL_PATH} or local, uncommitted zip at {ZIP_PATH}."
        )

    if RAW_DIR.exists():
        shutil.rmtree(RAW_DIR)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(str(ZIP_PATH), "r") as zf:
        for name in zf.namelist():
            if name.endswith(".xlsx"):
                zf.extract(name, str(RAW_DIR))
                # 寻找提取后的 xlsx 文件
                extracted = list(RAW_DIR.rglob("*.xlsx"))
                if extracted:
                    src = extracted[0]
                    if src != EXCEL_PATH:
                        shutil.move(str(src), str(EXCEL_PATH))
                    # 清理 zip 解压留下的空目录
                    for subdir in list(RAW_DIR.iterdir()):
                        if subdir.is_dir():
                            shutil.rmtree(subdir)
                    return EXCEL_PATH

    raise FileNotFoundError(f"在 zip {ZIP_PATH} 中未找到 .xlsx 文件")

# ═══════════════════════════════════════════════════════════════════════════
# 2. 读取 Excel
# ═══════════════════════════════════════════════════════════════════════════
def load_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    """返回 sheet 名称列表（按 Excel 原始顺序）。"""
    return wb.sheetnames


def read_queries(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """读取查询 sheet（训练集/测试集/验证集/全部数据）。

    列映射:
      编号        → query_id
      用户问题    → query
      正确答案（模型名称） → gold_model_name
      数据集类型  → split
    """
    ws = wb[sheet_name]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        record = {
            "query_id": str(ws.cell(r, 1).value or "").strip(),
            "query": str(ws.cell(r, 2).value or "").strip(),
            "gold_model_name": str(ws.cell(r, 3).value or "").strip(),
            "split": str(ws.cell(r, 4).value or "").strip(),
            "raw_row_index": r,
        }
        if record["query_id"]:
            # 编号可能为 None 或空行
            if not record["query_id"].startswith(("train_", "test_", "val_")):
                # 可能是全行空白，跳过
                if not any(v for v in record.values() if v):
                    continue
            rows.append(record)
    return rows


def read_models(wb: openpyxl.Workbook) -> list[dict]:
    """读取模型清单_参考。

    列映射:
      序号        → original_index
      业务模型全称 → model_name
      业务模型概述 → description
    """
    ws = wb[SHEET_MODELS]
    models: list[dict] = []
    for r in range(2, ws.max_row + 1):
        idx = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        if idx is None and not name:
            continue
        models.append({
            "original_index": int(idx) if idx is not None else None,
            "model_name": str(name or "").strip(),
            "description": str(desc or "").strip(),
        })
    return models


def read_stats(wb: openpyxl.Workbook) -> list[dict]:
    """读取统计信息。

    列映射:
      模型名称 → model_name
      总问题数 → total_count
      训练集   → train_count
      测试集   → test_count
      验证集   → val_count
    """
    ws = wb[SHEET_STATS]
    stats: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        total = ws.cell(r, 2).value
        train = ws.cell(r, 3).value
        test = ws.cell(r, 4).value
        val = ws.cell(r, 5).value
        if not name:
            continue
        stats.append({
            "model_name": str(name).strip(),
            "total_count": int(total) if total is not None else 0,
            "train_count": int(train) if train is not None else 0,
            "test_count": int(test) if test is not None else 0,
            "val_count": int(val) if val is not None else 0,
        })
    return stats


# ═══════════════════════════════════════════════════════════════════════════
# 3. 模型去重与 ID 分配
# ═══════════════════════════════════════════════════════════════════════════
def _normalize_model_name(name: str) -> str:
    """标准化模型名称：去首尾空白、统一内部空白。"""
    import re
    return re.sub(r"\s+", " ", name).strip()


def dedup_models(raw_models: list[dict]) -> tuple[list[dict], list[dict]]:
    """对原始模型列表去重，返回 (deduped_models, duplicate_info)。

    按标准化后的 model_name 去重，保留 original_index 最小的记录。
    分配 model_id 为 OFFICIAL_001 ~ OFFICIAL_060，按 original_index 升序。
    """
    # 按 original_index 排序
    sorted_models = sorted(raw_models, key=lambda m: m["original_index"] or 0)

    seen: dict[str, list[dict]] = {}
    for m in sorted_models:
        key = _normalize_model_name(m["model_name"])
        seen.setdefault(key, []).append(m)

    deduped: list[dict] = []
    duplicates: list[dict] = []

    for name_key, entries in seen.items():
        if len(entries) > 1:
            # 保留 index 最小的
            entries.sort(key=lambda e: e["original_index"] or 999)
            kept = entries[0]
            for dropped in entries[1:]:
                duplicates.append({
                    "model_name": name_key,
                    "kept_original_index": kept["original_index"],
                    "dropped_original_index": dropped["original_index"],
                })
            deduped.append(kept)
        else:
            deduped.append(entries[0])

    # 按 original_index 升序重新排序，分配 OFFICIAL_ID
    deduped.sort(key=lambda m: m["original_index"] or 0)
    for i, m in enumerate(deduped, start=1):
        m["model_id"] = f"OFFICIAL_{i:03d}"

    return deduped, duplicates


# ═══════════════════════════════════════════════════════════════════════════
# 4. 模型元数据启发式
# ═══════════════════════════════════════════════════════════════════════════
def _infer_domain(model_name: str, description: str) -> str:
    """从 model_name 和 description 关键词推断 domain。"""
    text = model_name + " " + description
    for keywords, domain in DOMAIN_KEYWORDS:
        if any(kw in text for kw in keywords):
            return domain
    return DOMAIN_FALLBACK


def _infer_business_scenario(model_name: str, description: str) -> str:
    """从模型名称或描述提取业务场景，控制在 20 字内。"""
    # 尝试从 model_name 去掉"模型"后缀取核心
    name = model_name
    for suffix in ["模型", "模型（", "模型("]:
        if suffix in name:
            name = name.split(suffix)[0]
            break
    name = name.strip()
    if len(name) <= 20 and len(name) > 2:
        return name

    # fallback: 从 description 第一句提取
    first_line = description.split("\n")[0].strip()
    # 去掉 "目标用户:" / "用途:" 等前缀
    for prefix in ["目标用户", "目标客户", "用途", "目的", "定义"]:
        if prefix in first_line:
            first_line = first_line.split(prefix, 1)[-1]
            if "：" in first_line:
                first_line = first_line.split("：", 1)[-1]
            if ":" in first_line:
                first_line = first_line.split(":", 1)[-1]
            break
    first_line = first_line.strip().rstrip("。，,.")
    if len(first_line) > 20:
        first_line = first_line[:20]
    if len(first_line) > 2:
        return first_line
    return model_name[:20]


def _extract_tags(model_name: str, description: str) -> list[str]:
    """从 model_name + description 提取 3-8 个业务关键词。"""
    import re
    text = model_name + " " + description
    # 按中文/英文/数字分词
    tokens: list[str] = []
    # 匹配中文字符序列或英文单词或数字
    for part in re.findall(r"[\u4e00-\u9fff]+|[A-Za-z0-9]+", text):
        part = part.strip()
        if not part:
            continue
        if part.lower() in STOP_WORDS:
            continue
        # 过滤单个字的 token（纯中文单字无意义）
        if re.match(r"^[\u4e00-\u9fff]$", part):
            continue
        tokens.append(part.lower())

    # 去重，保持顺序
    seen: set[str] = set()
    result: list[str] = []
    for t in tokens:
        t_lower = t.lower()
        if t_lower not in seen:
            seen.add(t_lower)
            result.append(t_lower)
        if len(result) >= 8:
            break

    if len(result) < 3:
        # 不足 3 个时从 model_name 补充
        extra = re.findall(r"[\u4e00-\u9fff]{2,}", model_name)
        for ex in extra:
            ex_lower = ex.lower()
            if ex_lower not in seen:
                seen.add(ex_lower)
                result.append(ex_lower)
            if len(result) >= 3:
                break

    # 限制最少 3 个
    while len(result) < 3:
        for suffix in ["银行", "模型", "业务"]:
            if suffix not in result:
                result.append(suffix)
            if len(result) >= 3:
                break

    return result


def build_model_enriched(deduped_models: list[dict]) -> list[dict]:
    """为去重后的模型补充 domain / business_scenario / tags / source_type。"""
    enriched: list[dict] = []
    for m in deduped_models:
        domain = _infer_domain(m["model_name"], m["description"])
        scenario = _infer_business_scenario(m["model_name"], m["description"])
        tags = _extract_tags(m["model_name"], m["description"])
        enriched.append({
            "model_id": m["model_id"],
            "original_index": m["original_index"],
            "model_name": m["model_name"],
            "description": m["description"],
            "domain": domain,
            "business_scenario": scenario,
            "tags": tags,
            "source_type": "official_dataset",
        })
    return enriched


# ═══════════════════════════════════════════════════════════════════════════
# 5. 构建查询文件
# ═══════════════════════════════════════════════════════════════════════════
def _build_model_name_to_id_map(enriched_models: list[dict]) -> dict[str, str]:
    """建立标准化 model_name → model_id 映射。"""
    mapping: dict[str, str] = {}
    for m in enriched_models:
        key = _normalize_model_name(m["model_name"])
        mapping[key] = m["model_id"]
    return mapping


def build_query_records(
    queries: list[dict],
    name_to_id: dict[str, str],
    split_label: str | None = None,
) -> list[dict]:
    """构建查询 JSONL 记录，将 gold_model_name 解析为 gold_model_ids。

    若 split_label 为 None，则使用记录中的 split 字段。
    """
    records: list[dict] = []
    unresolved: set[str] = set()
    for q in queries:
        split = split_label if split_label else q["split"]
        model_name_normalized = _normalize_model_name(q["gold_model_name"])
        model_id = name_to_id.get(model_name_normalized)
        gold_model_names = [q["gold_model_name"]]

        if not model_id:
            unresolved.add(q["gold_model_name"])
            continue

        records.append({
            "query_id": q["query_id"],
            "query": q["query"],
            "split": split,
            "gold_model_ids": [model_id],
            "gold_model_names": gold_model_names,
            "source_type": "official_dataset",
            "raw_row_index": q["raw_row_index"],
        })

    if unresolved:
        print(f"  [错误] 以下 gold_model_name 无法映射到任何 OFFICIAL_ID:")
        for name in sorted(unresolved):
            print(f"    - {repr(name)}")
        sys.exit(1)

    return records


# ═══════════════════════════════════════════════════════════════════════════
# 6. 报告生成
# ═══════════════════════════════════════════════════════════════════════════
def generate_inspection_report(
    excel_path: Path,
    sheet_names: list[str],
    sheet_row_counts: dict[str, int],
    field_mapping: dict,
    train_count: int,
    test_count: int,
    val_count: int,
    raw_model_count: int,
    deduped_model_count: int,
    duplicate_model_names: list[str],
) -> tuple[dict, str]:
    """生成 official_excel_inspection 报告。"""
    report = {
        "zip_path": project_relative(ZIP_PATH),
        "zip_committed": False,
        "excel_file": project_relative(excel_path),
        "sheet_names": sheet_names,
        "sheet_row_counts": sheet_row_counts,
        "field_mapping": field_mapping,
        "train_count": train_count,
        "test_count": test_count,
        "val_count": val_count,
        "raw_model_rows": raw_model_count,
        "deduped_model_count": deduped_model_count,
        "duplicate_model_names": duplicate_model_names,
    }
    md_lines = [
        "# 官方 Excel 检查报告",
        "",
        f"- **Zip 路径**: `{project_relative(ZIP_PATH)}`（本地原始压缩包，不随仓库提交）",
        f"- **Excel 路径**: `{project_relative(excel_path)}`",
        "",
        "## Sheet 列表",
        "",
        "| Sheet | 行数（含表头） | 数据行数 |",
        "|-------|---------------|---------|",
    ]
    for sn in sheet_names:
        total = sheet_row_counts.get(sn, 0)
        data_rows = total - 1 if total > 0 else 0
        md_lines.append(f"| {sn} | {total} | {data_rows} |")

    md_lines += [
        "",
        "## 字段映射",
        "",
    ]
    for sheet, mapping in field_mapping.items():
        md_lines.append(f"### {sheet}")
        for orig, mapped in mapping.items():
            md_lines.append(f"- `{orig}` → `{mapped}`")
        md_lines.append("")

    md_lines += [
        "## 数据集划分",
        "",
        f"- 训练集: {train_count}",
        f"- 测试集: {test_count}",
        f"- 验证集: {val_count}",
        f"- 全部数据: {train_count + test_count + val_count}",
        "",
        "## 模型清单",
        "",
        f"- 原始模型行数: {raw_model_count}",
        f"- 去重后模型数: {deduped_model_count}",
        f"- 重复模型名称: {duplicate_model_names or '无'}",
        "",
    ]

    return report, "\n".join(md_lines)


def generate_duplicate_report(duplicates: list[dict]) -> tuple[dict, str]:
    """生成 duplicate_models 报告。"""
    report = {
        "total_duplicate_rows": len(duplicates),
        "details": duplicates,
        "dedup_reason": (
            "小微企业违约概率B卡模型在序号 8 和 44 重复出现，描述完全相同。"
            "保留序号较小的（8），序号 44 被合并。"
            "去重后模型总数严格为 60。"
        ),
    }
    md_lines = [
        "# 重复模型报告",
        "",
        f"**总重复行数**: {len(duplicates)}",
        "",
        "| 模型名称 | 保留序号 | 合并序号 |",
        "|---------|---------|---------|",
    ]
    for d in duplicates:
        md_lines.append(
            f"| {d['model_name']} | {d['kept_original_index']} | "
            f"{d['dropped_original_index']} |"
        )
    md_lines += [
        "",
        "## 去重原因",
        "",
        report["dedup_reason"],
        "",
        "### 影响",
        "",
        "- 原始模型清单 61 行 → 去重后 60 个唯一模型",
        "- 模型 ID 使用 `OFFICIAL_001` 到 `OFFICIAL_060`",
        "- 重复项的查询仍映射到保留的 OFFICIAL_ID",
        "",
    ]
    return report, "\n".join(md_lines)


def generate_distribution_report(
    enriched_models: list[dict],
    stats_data: list[dict],
) -> tuple[dict, str]:
    """生成 model_question_distribution 报告。"""
    # 构建 stats 查找
    stats_map: dict[str, dict] = {}
    for s in stats_data:
        key = _normalize_model_name(s["model_name"])
        stats_map[key] = s

    distribution: list[dict] = []
    count6_models: list[str] = []
    count0_models: list[str] = []
    for m in enriched_models:
        key = _normalize_model_name(m["model_name"])
        st = stats_map.get(key, {})
        total = st.get("total_count", 0)
        train = st.get("train_count", 0)
        test = st.get("test_count", 0)
        val = st.get("val_count", 0)
        entry = {
            "model_id": m["model_id"],
            "model_name": m["model_name"],
            "total_count": total,
            "train_count": train,
            "test_count": test,
            "val_count": val,
        }
        distribution.append(entry)
        if total == 6:
            count6_models.append(m["model_name"])
        if total == 0:
            count0_models.append(m["model_name"])

    distribution.sort(key=lambda x: (x["total_count"], x["model_id"]))

    total7_count = sum(1 for d in distribution if d["total_count"] == 7)
    total6_count = len(count6_models)
    total0_count = len(count0_models)

    report = {
        "distribution": distribution,
        "summary": {
            "total_models": len(distribution),
            "models_with_7_questions": total7_count,
            "models_with_6_questions": total6_count,
            "models_with_0_questions": total0_count,
            "count6_model_names": count6_models,
            "conclusion": (
                f"{total7_count} 个模型 total_count=7，"
                f"{total6_count} 个模型 total_count=6，"
                f"{total0_count} 个模型 total_count=0，无明显高频异常"
            ),
        },
    }

    md_lines = [
        "# 模型问题数量分布报告",
        "",
        "## 统计结论",
        "",
        report["summary"]["conclusion"],
        "",
        f"### 详细分布",
        "",
        "| model_id | model_name | total_count | train_count | test_count | val_count |",
        "|----------|-----------|-------------|-------------|------------|-----------|",
    ]
    for d in distribution:
        md_lines.append(
            f"| {d['model_id']} | {d['model_name']} | {d['total_count']} | "
            f"{d['train_count']} | {d['test_count']} | {d['val_count']} |"
        )

    if count6_models:
        md_lines += [
            "",
            "### total_count=6 的模型",
            "",
        ]
        for name in count6_models:
            md_lines.append(f"- {name}")

    return report, "\n".join(md_lines)


# ═══════════════════════════════════════════════════════════════════════════
# 7. 写文件工具函数
# ═══════════════════════════════════════════════════════════════════════════
def write_json(path: Path, data) -> None:
    _ensure_dir(path.parent)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  OK: {path.name} ({len(json.dumps(data, ensure_ascii=False))} bytes)")


def write_jsonl(path: Path, records: list[dict]) -> None:
    _ensure_dir(path.parent)
    with open(path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"  OK: {path.name} ({len(records)} lines)")


def write_md(path: Path, content: str) -> None:
    _ensure_dir(path.parent)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  OK: {path.name} ({len(content)} bytes)")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════
def main() -> None:
    print("=" * 60)
    print("prepare_official_dataset.py")
    print("=" * 60)

    # ── 1. 准备 Excel ──────────────────────────────────────────────────
    print("\n[1/7] 准备 Excel...")
    excel_path = extract_excel()
    if ZIP_PATH.exists():
        print(f"  本地 zip: {project_relative(ZIP_PATH)}")
    else:
        print("  本地 zip: 未提交/未提供，使用已提交 Excel")
    print(f"  Excel: {project_relative(excel_path)}")
    # ── 2. 读取 Excel ──────────────────────────────────────────────────
    print("\n[2/7] 读取 Excel...")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    sheet_names = load_sheet_names(wb)
    print(f"  Sheets: {sheet_names}")

    # 读取各 sheet
    train_queries = read_queries(wb, SHEET_TRAIN)
    test_queries = read_queries(wb, SHEET_TEST)
    val_queries = read_queries(wb, SHEET_VAL)
    all_queries = read_queries(wb, SHEET_ALL)
    raw_models = read_models(wb)
    stats_data = read_stats(wb)

    sheet_row_counts: dict[str, int] = {}
    for sn in sheet_names:
        sheet_row_counts[sn] = wb[sn].max_row

    print(f"  训练集: {len(train_queries)}")
    print(f"  测试集: {len(test_queries)}")
    print(f"  验证集: {len(val_queries)}")
    print(f"  全部数据: {len(all_queries)}")
    print(f"  模型清单: {len(raw_models)}")
    print(f"  统计信息: {len(stats_data)}")

    # 验证计数
    assert len(train_queries) == 291, f"训练集应为 291，但为 {len(train_queries)}"
    assert len(test_queries) == 62, f"测试集应为 62，但为 {len(test_queries)}"
    assert len(val_queries) == 64, f"验证集应为 64，但为 {len(val_queries)}"
    assert len(all_queries) == 417, f"全部数据应为 417，但为 {len(all_queries)}"
    assert len(raw_models) == 61, f"模型清单应为 61，但为 {len(raw_models)}"
    assert len(stats_data) == 60, f"统计信息应为 60，但为 {len(stats_data)}"
    print("  [OK] 数量验证通过")

    # ── 3. 模型去重 ────────────────────────────────────────────────────
    print("\n[3/7] 模型去重...")
    deduped_models, duplicates = dedup_models(raw_models)
    print(f"  原始模型数: {len(raw_models)}")
    print(f"  去重后模型数: {len(deduped_models)}")
    assert len(deduped_models) == 60, f"去重后应为 60，但为 {len(deduped_models)}"
    for d in duplicates:
        print(f"  重复: {d['model_name']} (保留 {d['kept_original_index']}, "
              f"合并 {d['dropped_original_index']})")
    print("  [OK] 去重验证通过")

    # ── 4. 构建 enriched 模型元数据 ────────────────────────────────────
    print("\n[4/7] 构建模型元数据...")
    enriched_models = build_model_enriched(deduped_models)
    name_to_id = _build_model_name_to_id_map(enriched_models)
    print(f"  {len(enriched_models)} models with OFFICIAL_xxx IDs")

    # ── 5. 构建查询记录 ────────────────────────────────────────────────
    print("\n[5/7] 构建查询记录...")
    all_records = build_query_records(all_queries, name_to_id, split_label=None)
    train_records = build_query_records(train_queries, name_to_id, split_label="train")
    test_records = build_query_records(test_queries, name_to_id, split_label="test")
    val_records = build_query_records(val_queries, name_to_id, split_label="val")
    print(f"  query_model_eval.jsonl: {len(all_records)}")
    print(f"  queries_train.jsonl: {len(train_records)}")
    print(f"  queries_test.jsonl: {len(test_records)}")
    print(f"  queries_val.jsonl: {len(val_records)}")
    assert len(all_records) == 417
    assert len(train_records) == 291
    assert len(test_records) == 62
    assert len(val_records) == 64
    print("  [OK] 查询数量验证通过")

    # ── 6. 写数据文件 ──────────────────────────────────────────────────
    print("\n[6/7] 写数据文件...")

    # models.jsonl
    write_jsonl(OFFICIAL_DIR / "models.jsonl", enriched_models)

    # models_official_60.json
    write_json(OFFICIAL_DIR / "models_official_60.json", enriched_models)

    # query_model_eval.jsonl
    write_jsonl(OFFICIAL_DIR / "query_model_eval.jsonl", all_records)

    # queries_train.jsonl
    write_jsonl(OFFICIAL_DIR / "queries_train.jsonl", train_records)

    # queries_test.jsonl
    write_jsonl(OFFICIAL_DIR / "queries_test.jsonl", test_records)

    # queries_val.jsonl
    write_jsonl(OFFICIAL_DIR / "queries_val.jsonl", val_records)

    # dataset_manifest.json
    manifest = {
        "source": "user_excel_dataset",
        "source_archive": project_relative(ZIP_PATH),
        "source_archive_committed": False,
        "excel_file": project_relative(excel_path),
        "model_count": 60,
        "raw_model_rows": 61,
        "duplicate_model_rows": 1,
        "query_count": 417,
        "split_counts": {
            "train": 291,
            "test": 62,
            "val": 64,
        },
        "source_type_values": ["official_dataset"],
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    write_json(OFFICIAL_DIR / "dataset_manifest.json", manifest)

    # ── 7. 写报告文件 ──────────────────────────────────────────────────
    print("\n[7/7] 写报告文件...")

    field_mapping = {
        "训练集/测试集/验证集/全部数据": {
            "编号": "query_id",
            "用户问题": "query",
            "正确答案（模型名称）": "gold_model_name",
            "数据集类型": "split",
        },
        "模型清单_参考": {
            "序号": "original_index",
            "业务模型全称": "model_name",
            "业务模型概述": "description",
        },
        "统计信息": {
            "模型名称": "model_name",
            "总问题数": "total_count",
            "训练集": "train_count",
            "测试集": "test_count",
            "验证集": "val_count",
        },
    }

    duplicate_names = list(set(d["model_name"] for d in duplicates))

    ins_report, ins_md = generate_inspection_report(
        excel_path, sheet_names, sheet_row_counts, field_mapping,
        len(train_queries), len(test_queries), len(val_queries),
        len(raw_models), len(deduped_models), duplicate_names,
    )
    write_json(REPORT_DIR / "official_excel_inspection.json", ins_report)
    write_md(REPORT_DIR / "official_excel_inspection.md", ins_md)

    dup_report, dup_md = generate_duplicate_report(duplicates)
    write_json(REPORT_DIR / "duplicate_models.json", dup_report)
    write_md(REPORT_DIR / "duplicate_models.md", dup_md)

    dist_report, dist_md = generate_distribution_report(enriched_models, stats_data)
    write_json(REPORT_DIR / "model_question_distribution.json", dist_report)
    write_md(REPORT_DIR / "model_question_distribution.md", dist_md)

    # ── 汇总 ──────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("完成!")
    print("=" * 60)
    print(f"  数据文件: {OFFICIAL_DIR}")
    print(f"  报告文件: {REPORT_DIR}")
    print(f"  Excel: {excel_path}")
    print(f"  模型数: 60 (去重后, 原 61)")
    print(f"  查询总数: 417 (训练 {291}, 测试 {62}, 验证 {64})")


if __name__ == "__main__":
    main()
